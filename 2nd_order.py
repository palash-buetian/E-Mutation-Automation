# coding=utf-8


import configparser
import csv
import os
import re
import socket
import time
from datetime import date
import datetime

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait

# parse configuration file
config = configparser.ConfigParser()
config.read('config.ini', encoding='utf-8')


# print(chr(27) + "[2J")

def is_connected(hostname):
    try:
        # see if we can resolve the host name -- tells us if there is
        # a DNS listening
        host = socket.gethostbyname(hostname)
        # connect to the host -- tells us if the host is actually
        # reachable
        s = socket.create_connection((host, 80), 2)
        s.close()
        return True
    except:
        pass
    return True


def credentials():
    # get username and password from config file
    username = config.get('LOGIN', 'username')
    password = config.get('LOGIN', 'password')
    return {'username': username, 'password': password}


def prepare_browser():
    # create a browser_instance instance with infobar disabled and password manager disabled
    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_experimental_option("excludeSwitches", ['enable-automation', 'enable-logging'])
    prefs = {"credentials_enable_service": False, "profile.password_manager_enabled": False}
    chrome_options.add_experimental_option("prefs", prefs)
    browser = webdriver.Chrome(ChromeDriverManager().install(), options=chrome_options)
    os.system('cls' if os.name == 'nt' else 'clear')
    # maximizing the browser_instance window make all the links visible.
    browser.maximize_window()
    return browser


def close_session():
    browser.stop_client()
    browser.quit()


def wait_and_click_by_xpath(wait, xpath):
    button = wait.until(EC.element_to_be_clickable((By.XPATH, xpath)))
    # click the send send button finally
    browser.execute_script("arguments[0].click();", button)


def login(browser, credentials):
    # land on login page
    login_url = config.get('URLS', 'login_page')
    welcome_page = config.get('URLS', 'welcome_page')

    browser.get(login_url)

    username = credentials['username']
    password = credentials['password']

    # try login
    browser.find_element_by_id('username').send_keys(username)
    browser.find_element_by_id('password').send_keys(password)    
    wait_and_click_by_xpath(wait, "//*[@id='loginBtn']")
    current_url = browser.current_url
    WebDriverWait(browser, 15).until(EC.url_changes(current_url))
    
    # if login unsuccessful
    if browser.current_url != welcome_page:
        print('login error! ')
        close_session()
        #browser.get(welcome_page)
        return True
    else:
        print("Login successful! ")
        return True

def click_by_xpath(xpath):
    element = browser.find_element_by_xpath(xpath)
    browser.execute_script("arguments[0].click();", element)

def click_by_linktext(linktext):
    element = browser.find_element_by_link_text(linktext)
    browser.execute_script("arguments[0].click();", element)


def proposal_case_found(proposal_url, processed=False):
    # open the acceptable application list page

    browser.get(proposal_url)

    cases = len(browser.find_elements_by_xpath("//*[@id='sample_6']/tbody/tr"))
    # if there is no case, end here.

    if cases == 0:
        if processed:
            print("All 2nd order processed. Job done! ")
        else:
            print("No 2nd order pending case found! ")
        return False

    return True


def initiate_report_file():
    # the name of output csv file as 'current date'
    order_datetime = date.today().strftime("%d-%m-%Y") + '_' + time.strftime("%I.%m-%p", time.localtime())
    folder_location = config.get('REPORT', '2nd_order_report_folder_path')
    file_name = folder_location + order_datetime + '_2nd_Order'
    return file_name


def initiate_csv_writer():
    folder_path = config.get('REPORT', '2nd_order_report_folder_path')
    # if folder don't exist, create one
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)


def write_header_column(file_name):
    # create/open if exist, the csv file to read+write
    with open(file_name + '.csv', mode='a+', newline='', encoding='utf-8') as f:
        writer = csv.writer(f, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
        # write the column names of the csv file if it is not already written
        if os.stat(file_name + '.csv').st_size == 0:
            writer.writerow(
                ['আবেদন নং', 'নামজারি মামলা নং', 'শুনানির তারিখ', 'আবেদনকারীর নাম', 'দাখিলের তারিখ', 'মৌজা'])


def application_details():
    details = {}
    # get the basic information from the table
    details['application_number'] = browser.find_element_by_xpath(
        "//*[@id='sample_6']/tbody/tr[1]/td[1]").text
    details['application_filling_date'] = browser.find_element_by_xpath(
        "//*[@id='sample_6']/tbody/tr[1]/td[2]").text
    details['applicant_name'] = browser.find_element_by_xpath("//*[@id='sample_6']/tbody/tr[1]/td[5]").text
    details['mouza'] = browser.find_element_by_xpath("//*[@id='sample_6']/tbody/tr[1]/td[6]").text
    details['application_details_page_url'] = browser.find_element_by_link_text(
        "পদক্ষেপ নিন").get_attribute('href')

    return details


def second_order_exists():
    try:
        browser.find_element_by_xpath("//*[@id='order_data']/tr[2]/td[2]")
        return True
    except NoSuchElementException:
        return False


def date_by_adding_business_days(from_date, add_days):
    business_days_to_add = add_days + 1
    current_date = from_date
    while business_days_to_add > 0:
        current_date += datetime.timedelta(days=1)
        weekday = current_date.isoweekday()
        if (weekday == 5) or (weekday == 6):  # friday = 5, saterday = 6
            continue
        business_days_to_add -= 1
    return current_date


def en2bn(number):
    edit = [("১", "1"), ("২", "2"), ("৩", "3"), ("৪", "4"), ("৫", "5"), ("৬", "6"), ("৭", "7"), ("৮", "8"),
            ("৯", "9"), ("০", "0")]
    for search, replace in edit:
        number = str(number).replace(replace, search)
    return number


def prepare_hearing_date(increment=0):
    interval = int(config.get("SECOND_ORDER", "second_order_interval_days"))
    # prepare sencond order
    hearing_date = date_by_adding_business_days(datetime.date.today(), interval + increment)
    return hearing_date


def add_second_order(hearing_date):
    # # click 'আদেশ সংযুক্তকরণ' link to add 1st order
    wait_and_click_by_xpath(wait, "//*[@id='add_order']")

    # # Sleep for 1s
    time.sleep(2)

    # # select the order with "কানুনগো/সার্ভেয়রের প্রতিবেদন ও শুনানির জন্য আদেশ", option id =23
    select = Select(browser.find_element_by_id('status-id'))
    select.select_by_value('23')
    # change hearing date by js
    js = '$(".date-picker").datepicker("setDate",' + f'"{hearing_date}"' + ');'
    browser.execute_script(js)
    # # to load the full order on ckeditor, wait three seconds
    time.sleep(2)

    # now click 'সংরক্ষণ' button to save the 1st order
    click_by_xpath("//*[@id='submit_order_btn']/button")

    # sleep for 2s
    time.sleep(1)

def check_order_signature():
    try:
        browser.find_element_by_xpath(
            "//*[@id='sample_6']/tbody/tr[1]/td[1]/div[1]/table/tbody[2]/tr[1]/td[2]/div[2]/a")
        return False
    except NoSuchElementException:
        return True


def check_notice_signature():
    # click the notice tab on right side
    button = wait.until(EC.element_to_be_clickable((By.LINK_TEXT, "নোটিশ/শুনানি")))
    # click the send send button finally
    browser.execute_script("arguments[0].click();", button)

    # click the notice full page link within the tab
    wait_and_click_by_xpath(wait, "//*[@id='view_notice_link_id']")
    time.sleep(2.5)
    try:
        browser.find_element_by_xpath(
            "//*[@id='portlet_tab5']/div[2]/div/div/table/tbody/tr[10]/td[3]/div[1]/a")
        return False
    except NoSuchElementException:
        return True


def put_notice_signature():
    # signature modal for notice link xpath
    wait_and_click_by_xpath(wait, "//*[@id='portlet_tab5']/div[2]/div/div/table/tbody/tr[10]/td[3]/div[1]/a")

    # on the notice signature popup modal, first get the 'স্বাক্ষর প্রদান করুন' button
    wait_and_click_by_xpath(wait, "//*[@id='signature']/form/div[2]/button")

    # click the notice full page link within the tab
    wait_and_click_by_xpath(wait, "//*[@id='view_notice_link_id']")

    # click the send notification sms button
    wait_and_click_by_xpath(wait, "//*[@id='portlet_tab5']/div[4]/button[4]")


def put_order_signature():
    # signature modal link click
    click_by_xpath("//*[@id='sample_6']/tbody/tr[1]/td[1]/div[1]/table/tbody[2]/tr[1]/td[2]/div[2]/a")

    # on the signature popup modal, first get the 'স্বাক্ষর প্রদান করুন' button
    sign_button = wait.until(
        EC.element_to_be_clickable((By.XPATH, "//*[@id='signature']/form/div[2]/button")))
    # Click the 'স্বাক্ষর প্রদান করুন' button
    sign_button.click()


def send():
    # open send popup modal
    popup_button = browser.find_element_by_css_selector(
        "body > div.page-container > div.page-content-wrapper > div > div > div > div > div.portlet-title > div.actions > button")
    browser.execute_script("arguments[0].click();", popup_button)

    time.sleep(1)
    browser.find_element_by_id("users-tagged-receive-7").click()
    time.sleep(0.5)
    browser.find_element_by_id("users-tagged-view-3").click()

    # Send button xpath
    send_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//*[@id='submit-btn']")))
    # click the send send button finally
    send_button.click()


def convert_csv_to_xlsx(file):
    # Convert csv to xlsx
    wb = Workbook()
    ws = wb.active
    with open(file + '.csv', 'r', encoding="utf-8") as f:
        for row in csv.reader(f): ws.append(row)
    wb.save(file + '.xlsx')


def xlsx_formating(ws):
    # insert a new column for making serial
    ws.insert_cols(1)
    ws['A1'] = 'ক্রমিক'

    # insert value to the serial
    for cell in range(2, int(ws.max_row) + 1):
        # set 1,2,3,4 in serial column
        ws.cell(row=cell, column=1).value = cell - 1

    # format all spreadsheet with style
    for cell in range(1, int(ws.max_row) + 1):
        for col in range(1, int(ws.max_column) + 1):
            ws.cell(row=cell, column=col).alignment = Alignment(horizontal='center')
            ws.cell(row=cell, column=col).font = Font(name='NikoshBAN', size=18)
            ws.cell(row=cell, column=col).border = Border(left=Side(style='thin'),
                                                          right=Side(style='thin'),
                                                          top=Side(style='thin'),
                                                          bottom=Side(style='thin'))

    # Header row font style
    header_row_style = Font(bold=True, name='NikoshBAN', size=20)
    for cell in ws["1:1"]:
        cell.font = header_row_style

    # set the printing parameters
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = True
    ws.print_area = 'A1' + ':G' + str(ws.max_row) + ''
    ws.firstFooter.right.text = "Page &[Page] of &N"
    ws.set_printer_settings(paper_size=9, orientation='portrait')
    ws.page_setup.fitToPage = True

    # adjust width of the columns
    for col in ws.columns:
        max_lenght = 0
        col_name = re.findall('\w\d', str(col[0]))
        col_name = col_name[0]
        col_name = re.findall('\w', str(col_name))[0]
        for cell in col:
            try:
                if len(str(cell.value)) > max_lenght:
                    max_lenght = len(cell.value)
            except:
                pass
        adjusted_width = (max_lenght * 1.8)
        ws.column_dimensions[col_name].width = adjusted_width


def write_application_details(file_name, details):
    with open(file_name + '.csv', mode='a+', newline='', encoding='utf-8') as f:
        # write the csv column values
        csv.writer(f, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
        writer = csv.writer(f, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
        writer.writerow([
            details['application_number'],
            details['new_case_number'],
            details['next_date'],
            details['applicant_name'],
            details['application_filling_date'],
            details['mouza']])


def show_info_in_terminal(details):
    # Terminal output
    print("----------------------------")
    print("Application Number: " + details['application_number'])
    print("Application Filling Date: " + details['application_filling_date'])
    print("Applicant Name: " + details['applicant_name'])
    print("Mouza: " + details['mouza'])
    print("Case Number: " + details['new_case_number'])
    print("Hearing Date: " + details['next_date'])
    print("----------------------------")




def notice_exists():
    try:
        browser.find_element_by_xpath("//*[@id='view_notice_link_id']")
        return True
    except NoSuchElementException:
        return False


def add_notice():
    # notice creation and sending

    # click the notice tab on right side
    button = wait.until(EC.element_to_be_clickable((By.LINK_TEXT, "নোটিশ/শুনানি")))
    # click the send send button finally
    browser.execute_script("arguments[0].click();", button)

    # click the notice full page link within the tab
    # notice_detail_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//*[@id='view_notice_link_id']")))
    # browser_instance.execute_script("arguments[0].click();", notice_detail_button)

    time.sleep(1)

    # notice submit for the first time
    try:
        notice_button = browser.find_element_by_xpath("//*[@id='notice_submit_id']")
        browser.execute_script("arguments[0].click();", notice_button)

        # sleep for 1 sec
        time.sleep(1)
        browser.refresh()
        # click the notice tab on right side
        button = wait.until(EC.element_to_be_clickable((By.LINK_TEXT, "নোটিশ/শুনানি")))
        # click the send send button finally
        browser.execute_script("arguments[0].click();", button)
        wait_and_click_by_xpath(wait, "//*[@id='view_notice_link_id']")
    except NoSuchElementException:
        button = wait.until(EC.element_to_be_clickable((By.LINK_TEXT, "নোটিশ/শুনানি")))
        browser.execute_script("arguments[0].click();", button)
        wait_and_click_by_xpath(wait, "//*[@id='notice_submit_id']")


    # click the notice full page link within the tab
    # signature modal for notice link xpath
    #wait_and_click_by_xpath(wait, "//*[@id='portlet_tab5']/div[2]/div/div/table/tbody/tr[10]/td[3]/div[1]/a")

    # on the notice signature popup modal, first get the 'স্বাক্ষর প্রদান করুন' button
    #wait_and_click_by_xpath(wait, "//*[@id='signature']/form/div[2]/button")

    # click the notice full page link within the tab
    #wait_and_click_by_xpath(wait, "//*[@id='view_notice_link_id']")

    # click the send notification sms button
    wait_and_click_by_xpath(wait, "//*[@id='portlet_tab5']/div[4]/button[4]")
    # sleep for 1 sec
    time.sleep(1)
    # close the modal
    wait_and_click_by_xpath(wait, "//*[@id='send_email_sms']/div[2]/div/div[3]/button")


if is_connected("one.one.one.one"):

    browser = prepare_browser()

    # 10 seconds timeout to load any page before it shows the error message
    wait = WebDriverWait(browser, 10)


    credentials = credentials()

    if login(browser, credentials) is True:

        # Click নামজারি link
        wait_and_click_by_xpath(wait,"*//a[@data-s_id='2']/span")
        
        time.sleep(1)
        browser.refresh()

        #  applications returned from ULAO list as sorted by application id as decending
        proposal_list = config.get('URLS', 'proposal_list').split(",")

        # two different page of 2nd order
        for proposal_url in proposal_list:
            # if new case is found-

            if proposal_case_found(proposal_url, processed=False) is True:

                file = initiate_report_file()
                initiate_csv_writer()
                write_header_column(file)

                case_handled = 0
                while True:
                    if proposal_case_found(proposal_url, processed=True):
                        # new case found
                        details = application_details()

                        # Let's start working on single application
                        # go to the case details page
                        browser.get(details['application_details_page_url'])

                        # check if there is any 1st order already.
                        if second_order_exists() is False:
                            case_cluster = config.get("SECOND_ORDER", "case_number_per_cluster")
                            increment = int(case_handled / int(case_cluster))
                            case_handled += 1
                            # no second order, so add one.
                            hearing_date = prepare_hearing_date(increment=increment)
                            add_second_order(hearing_date)
                        else:
                            # second order exists already
                            # if there is no signature, put one.
                            if check_order_signature() is False:
                                put_order_signature()

                        if notice_exists() is False:
                            # no second order notice created
                            add_notice()
                        else:
                            # second order notice is already there
                            # check and put signature if not already done.

                            if check_notice_signature() is False:
                                put_notice_signature()

                        # get mutation case number
                        details['new_case_number'] = browser.find_element_by_xpath("//*[@id='case_change_no']").text
                        # get the next date
                        details['next_date'] = browser.find_element_by_name(
                            "case_status_update[next_status_date]").get_attribute(
                            "value")
                        write_application_details(file, details)
                        show_info_in_terminal(details)

                        # send
                        send()

                    else:
                        # no new case.
                        break
                # now repeat the job for the rest of the case waiting for second order

                # Reporting starts
                # convert csv to xlsx
                convert_csv_to_xlsx(file)

                # open xlsx
                wb = load_workbook(filename=file + '.xlsx')
                ws = wb.active
                # format xlsx
                xlsx_formating(ws)

                # save the settings on xl file
                wb.save(file + '.xlsx')

                # now delete the csv file
                os.remove(file + '.csv')

    # end it
    close_session()
else:
    print("No Internet! Fix it first.")